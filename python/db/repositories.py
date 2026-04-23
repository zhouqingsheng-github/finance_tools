"""
数据访问对象（Repository）
封装所有数据库 CRUD 操作
"""

import json
import logging
import time
import uuid
from typing import Optional

from db.database import DatabaseManager
logger = logging.getLogger('finance-tools.db')


class MerchantRepository:
    """商家配置数据访问"""
    
    def __init__(self, db_path: str):
        self.db_path = db_path
    
    def _get_db(self) -> DatabaseManager:
        from db.database import init_database
        return init_database(self.db_path)
    
    def list_all(self) -> list[dict]:
        """获取所有商家配置"""
        db = self._get_db()
        conn = db.get_connection()
        
        try:
            cursor = conn.execute("""
                SELECT id, name, url, login_url,
                       username, password,
                       username_selector, password_selector, submit_selector,
                       cookie_domains, headers, api_endpoints,
                       status, timeout, wait_after_login, max_retries,
                       last_login_at, credential_expires_at,
                       created_at, updated_at
                FROM merchants 
                WHERE is_deleted = 0 
                ORDER BY created_at DESC
            """)
            
            merchants = [DatabaseManager.dict_from_row(row) for row in cursor.fetchall()]
            
            # 解析 JSON 字段
            for m in merchants:
                try:
                    m['cookie_domains'] = json.loads(m.get('cookie_domains', '[]'))
                    m['headers'] = json.loads(m.get('headers', '{}'))
                except (json.JSONDecodeError, TypeError):
                    pass
            
            return merchants
        finally:
            conn.close()

    def list_active(self) -> list[dict]:
        """获取所有启用的商家"""
        all_merchants = self.list_all()
        return [m for m in all_merchants if m.get('status') == 'active']

    def get_by_id(self, merchant_id: str) -> Optional[dict]:
        """根据 ID 获取单个商家"""
        db = self._get_db()
        conn = db.get_connection()
        
        try:
            cursor = conn.execute(
                "SELECT * FROM merchants WHERE id = ? AND is_deleted = 0",
                (merchant_id,)
            )
            row = cursor.fetchone()
            
            if not row:
                return None
            
            merchant = DatabaseManager.dict_from_row(row)
            
            # 解析 JSON 字段
            try:
                merchant['cookie_domains'] = json.loads(merchant.get('cookie_domains', '[]'))
                merchant['headers'] = json.loads(merchant.get('headers', '{}'))
                merchant['api_endpoints'] = json.loads(merchant.get('api_endpoints', '[]'))
            except (json.JSONDecodeError, TypeError):
                pass
            
            return merchant
        finally:
            conn.close()

    def create(self, data: dict) -> dict:
        """创建新商家"""
        db = self._get_db()
        conn = db.get_connection()
        
        now = int(time.time())
        merchant_id = data.get('id') or str(uuid.uuid4())[:8]
        
        try:
            conn.execute("""
                INSERT INTO merchants (
                    id, name, url, login_url,
                    username, password,
                    username_selector, password_selector, submit_selector,
                    cookie_domains, headers, api_endpoints,
                    status, timeout, wait_after_login, max_retries,
                    created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                merchant_id,
                data.get('name', ''),
                data.get('url', ''),
                data.get('login_url', '') or data.get('url', ''),
                data.get('username', ''),
                data.get('password', ''),
                data.get('username_selector', '#username'),
                data.get('password_selector', '#password'),
                data.get('submit_selector', '.btn-login, #submitBtn, button[type="submit"]'),
                json.dumps(data.get('cookie_domains', [])),
                json.dumps(data.get('headers', {})),
                json.dumps(data.get('api_endpoints', [])),
                data.get('status', 'active'),
                data.get('timeout', 30),
                data.get('wait_after_login', 3),
                data.get('max_retries', 3),
                now, now
            ))
            conn.commit()
            
            logger.info(f"Merchant created: {merchant_id} - {data.get('name')}")
            return self.get_by_id(merchant_id)
        finally:
            conn.close()

    def update(self, merchant_id: str, data: dict) -> bool:
        """更新商家信息"""
        db = self._get_db()
        conn = db.get_connection()
        
        now = int(time.time())
        update_fields = []
        values = []
        
        allowed_fields = [
            'name', 'url', 'login_url',
            'username', 'password',
            'username_selector', 'password_selector', 'submit_selector',
            'cookie_domains', 'headers', 'api_endpoints',
            'status', 'timeout', 'wait_after_login', 'max_retries'
        ]
        
        for field in allowed_fields:
            if field in data:
                # JSON 字段需要序列化
                if field in ['cookie_domains', 'headers', 'api_endpoints']:
                    update_fields.append(f"{field} = ?")
                    values.append(json.dumps(data[field]))
                else:
                    update_fields.append(f"{field} = ?")
                    values.append(data[field])
        
        if not update_fields:
            return False
        
        update_fields.append("updated_at = ?")
        values.append(now)
        values.append(merchant_id)
        
        try:
            conn.execute(
                f"UPDATE merchants SET {', '.join(update_fields)} WHERE id = ?",
                tuple(values)
            )
            conn.commit()
            logger.info(f"Merchant updated: {merchant_id}")
            return True
        finally:
            conn.close()

    def delete(self, merchant_id: str) -> bool:
        """删除商家（软删除）"""
        db = self._get_db()
        conn = db.get_connection()
        
        try:
            conn.execute(
                "UPDATE merchants SET is_deleted = 1, updated_at = ? WHERE id = ?",
                (int(time.time()), merchant_id)
            )
            conn.commit()
            logger.info(f"Merchant deleted: {merchant_id}")
            return True
        finally:
            conn.close()

    def update_last_login(self, merchant_id: str, login_time: int, credential_expires_at: int = 0):
        """更新最后登录时间和凭证过期时间"""
        db = self._get_db()
        conn = db.get_connection()
        
        try:
            if credential_expires_at > 0:
                conn.execute(
                    "UPDATE merchants SET last_login_at = ?, credential_expires_at = ?, updated_at = ? WHERE id = ?",
                    (login_time, credential_expires_at, int(time.time()), merchant_id)
                )
            else:
                conn.execute(
                    "UPDATE merchants SET last_login_at = ?, updated_at = ? WHERE id = ?",
                    (login_time, int(time.time()), merchant_id)
                )
            conn.commit()
        finally:
            conn.close()


class CredentialRepository:
    """凭证数据访问"""
    
    def __init__(self, db_path: str):
        self.db_path = db_path

    def _get_db(self) -> DatabaseManager:
        from db.database import init_database
        return init_database(self.db_path)

    def get_by_merchant(self, merchant_id: str) -> Optional[dict]:
        """获取商家的最新凭证"""
        db = self._get_db()
        conn = db.get_connection()
        
        try:
            cursor = conn.execute(
                "SELECT * FROM credentials WHERE merchant_id = ? ORDER BY created_at DESC LIMIT 1",
                (merchant_id,)
            )
            row = cursor.fetchone()
            
            if not row:
                return None
                
            cred = DatabaseManager.dict_from_row(row)
            
            try:
                cred['cookie_domains'] = json.loads(cred.get('cookie_domains', '[]'))
            except (json.JSONDecodeError, TypeError):
                pass
            
            return cred
        finally:
            conn.close()

    def upsert(self, merchant_id: str, credential_data: dict) -> dict:
        """创建或更新凭证"""
        db = self._get_db()
        conn = db.get_connection()
        
        now = int(time.time())
        
        try:
            # 检查是否已存在
            existing = self.get_by_merchant(merchant_id)
            
            if existing:
                conn.execute("""
                    UPDATE credentials SET 
                        encrypted_cookies = ?,
                        encrypted_cookie_string = ?,
                        encrypted_storage_state = ?,
                        cookie_domains = ?,
                        expires_at = ?,
                        source_url = ?,
                        is_valid = 1,
                        updated_at = ?
                    WHERE merchant_id = ?
                """, (
                    credential_data.get('encrypted_cookies', ''),
                    credential_data.get('encrypted_cookie_string', ''),
                    credential_data.get('encrypted_storage_state', ''),
                    json.dumps(credential_data.get('cookie_domains', [])),
                    credential_data.get('expires_at', 0),
                    credential_data.get('source_url', ''),
                    now,
                    merchant_id
                ))
            else:
                conn.execute("""
                    INSERT INTO credentials (
                        merchant_id, encrypted_cookies, encrypted_cookie_string,
                        encrypted_storage_state, cookie_domains, expires_at,
                        source_url, is_valid, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    merchant_id,
                    credential_data.get('encrypted_cookies', ''),
                    credential_data.get('encrypted_cookie_string', ''),
                    credential_data.get('encrypted_storage_state', ''),
                    json.dumps(credential_data.get('cookie_domains', [])),
                    credential_data.get('expires_at', 0),
                    credential_data.get('source_url', ''),
                    1,
                    now, now
                ))
            
            conn.commit()
            logger.info(f"Credential saved for merchant {merchant_id}")
            return self.get_by_merchant(merchant_id)
        finally:
            conn.close()

    def invalidate(self, merchant_id: str):
        """标记凭证失效"""
        db = self._get_db()
        conn = db.get_connection()
        
        try:
            conn.execute(
                "UPDATE credentials SET is_valid = 0, updated_at = ? WHERE merchant_id = ?",
                (int(time.time()), merchant_id)
            )
            conn.commit()
        finally:
            conn.close()


class TaskRepository:
    """任务配置数据访问"""

    def __init__(self, db_path: str):
        self.db_path = db_path

    def _get_db(self) -> DatabaseManager:
        from db.database import init_database
        return init_database(self.db_path)

    def list_all(self) -> list[dict]:
        """获取所有任务"""
        db = self._get_db()
        conn = db.get_connection()
        try:
            cursor = conn.execute("""
                SELECT * FROM tasks
                WHERE is_deleted = 0
                ORDER BY created_at DESC
            """)
            tasks = [DatabaseManager.dict_from_row(row) for row in cursor.fetchall()]
            self._enrich_task_merchants(tasks)
            return tasks
        finally:
            conn.close()

    def get_by_id(self, task_id: str) -> Optional[dict]:
        """获取单个任务"""
        db = self._get_db()
        conn = db.get_connection()
        try:
            cursor = conn.execute("""
                SELECT * FROM tasks
                WHERE id = ? AND is_deleted = 0
            """, (task_id,))
            row = cursor.fetchone()
            if not row:
                return None
            task = DatabaseManager.dict_from_row(row)
            self._enrich_task_merchants([task])
            return task
        finally:
            conn.close()

    def _enrich_task_merchants(self, tasks: list[dict]):
        """为任务列表反序列化 JSON 字段并补充商家名称"""
        all_merchant_ids = set()
        for t in tasks:
            try:
                t['headers'] = json.loads(t.get('headers', '{}'))
                t['params'] = json.loads(t.get('params', '{}'))
                t['field_mapping'] = json.loads(t.get('field_mapping', '{}'))
                t['response_extract'] = json.loads(t.get('response_extract', '{}'))
                t['pagination'] = json.loads(t.get('pagination', '{}'))
                t['browser_config'] = json.loads(t.get('browser_config', '{}'))
                raw_ids = t.get('merchant_ids', '[]')
                if isinstance(raw_ids, str):
                    mids = json.loads(raw_ids)
                elif isinstance(raw_ids, list):
                    mids = raw_ids
                else:
                    mids = []
                t['_merchant_ids_list'] = mids
                all_merchant_ids.update(mids)
            except (json.JSONDecodeError, TypeError):
                t['_merchant_ids_list'] = []

        # 批量查商家名
        if all_merchant_ids:
            db2 = self._get_db()
            conn2 = db2.get_connection()
            try:
                placeholders = ','.join('?' * len(all_merchant_ids))
                cur = conn2.execute(f"SELECT id, name FROM merchants WHERE id IN ({placeholders})", tuple(all_merchant_ids))
                name_map = {r['id']: r['name'] for r in [DatabaseManager.dict_from_row(r) for r in cur.fetchall()]}
                for t in tasks:
                    names = [name_map.get(mid, '') for mid in t.get('_merchant_ids_list', [])]
                    t['merchant_name'] = ', '.join(names)
                    t['merchant_ids'] = t.get('_merchant_ids_list', [])
                    del t['_merchant_ids_list']
                conn2.close()
            except Exception:
                pass

    def list_by_merchant(self, merchant_id: str) -> list[dict]:
        """获取某个商家的所有任务（merchant_ids JSON 数组中包含该商家ID）"""
        db = self._get_db()
        conn = db.get_connection()
        try:
            cursor = conn.execute("""
                SELECT * FROM tasks
                WHERE merchant_ids LIKE ? AND is_deleted = 0
                ORDER BY created_at DESC
            """, (f'%"{merchant_id}"%',))
            tasks = [DatabaseManager.dict_from_row(row) for row in cursor.fetchall()]
            self._enrich_task_merchants(tasks)
            return tasks
        finally:
            conn.close()

    def create(self, data: dict) -> dict:
        """创建任务"""
        db = self._get_db()
        conn = db.get_connection()
        now = int(time.time())
        task_id = data.get('id') or str(uuid.uuid4())[:8]
        try:
            conn.execute("""
                INSERT INTO tasks (
                    id, name, merchant_ids, curl_command,
                    method, url, headers, params, body,
                    inject_credential, field_mapping,
                    response_extract, pagination,
                    cron_expression, task_type, browser_config,
                    status, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                task_id,
                data.get('name', ''),
                json.dumps(data.get('merchant_ids', []), ensure_ascii=False),
                data.get('curl_command', ''),
                data.get('method', 'GET'),
                data.get('url', ''),
                json.dumps(data.get('headers', {}), ensure_ascii=False),
                json.dumps(data.get('params', {}), ensure_ascii=False),
                data.get('body', ''),
                data.get('inject_credential', 1),
                json.dumps(data.get('field_mapping', {}), ensure_ascii=False),
                json.dumps(data.get('response_extract', {}), ensure_ascii=False),
                json.dumps(data.get('pagination', {}), ensure_ascii=False),
                data.get('cron_expression', ''),
                data.get('task_type', 'curl'),
                json.dumps(data.get('browser_config', {}), ensure_ascii=False),
                data.get('status', 'idle'),
                now, now
            ))
            conn.commit()
            logger.info(f"Task created: {task_id} - {data.get('name')}")
            return self.get_by_id(task_id)
        finally:
            conn.close()

    def update(self, task_id: str, data: dict) -> bool:
        """更新任务"""
        db = self._get_db()
        conn = db.get_connection()
        now = int(time.time())
        update_fields = []
        values = []

        allowed_fields = [
            'name', 'merchant_ids', 'curl_command',
            'method', 'url', 'headers', 'params', 'body',
            'inject_credential', 'field_mapping',
            'response_extract', 'pagination',
            'cron_expression', 'status', 'last_run_at', 'last_result',
            'task_type', 'browser_config'
        ]

        for field in allowed_fields:
            if field in data:
                if field in ('headers', 'params', 'field_mapping', 'response_extract', 'pagination', 'merchant_ids', 'browser_config'):
                    update_fields.append(f"{field} = ?")
                    values.append(json.dumps(data[field], ensure_ascii=False))
                else:
                    update_fields.append(f"{field} = ?")
                    values.append(data[field])

        if not update_fields:
            return False

        update_fields.append("updated_at = ?")
        values.append(now)
        values.append(task_id)

        try:
            conn.execute(
                f"UPDATE tasks SET {', '.join(update_fields)} WHERE id = ?",
                tuple(values)
            )
            conn.commit()
            logger.info(f"Task updated: {task_id}")
            return True
        finally:
            conn.close()

    def delete(self, task_id: str) -> bool:
        """删除任务（软删除）"""
        db = self._get_db()
        conn = db.get_connection()
        try:
            conn.execute(
                "UPDATE tasks SET is_deleted = 1, updated_at = ? WHERE id = ?",
                (int(time.time()), task_id)
            )
            conn.commit()
            logger.info(f"Task deleted: {task_id}")
            return True
        finally:
            conn.close()


class DataRepository:
    """采集数据访问"""
    
    def __init__(self, db_path: str):
        self.db_path = db_path

    @property
    def db_path_str(self) -> str:
        return self.db_path

    def _get_db(self) -> DatabaseManager:
        from db.database import init_database
        return init_database(self.db_path)

    def create(self, data: dict) -> int:
        """创建采集记录（每次任务执行产生一条结构化JSON记录）"""
        db = self._get_db()
        conn = db.get_connection()
        
        try:
            cursor = conn.execute("""
                INSERT INTO collected_data (task_id, merchant_id, merchant_name, raw_data, collected_at)
                VALUES (?, ?, ?, ?, ?)
            """, (
                data.get('task_id', ''),
                data.get('merchant_id', ''),
                data.get('merchant_name', ''),
                json.dumps(data.get('raw_data', {}), ensure_ascii=False),
                data.get('collected_at', int(time.time()))
            ))
            conn.commit()
            return cursor.lastrowid
        finally:
            conn.close()

    def list(self, filters: dict = {}, page: int = 1, page_size: int = 15) -> tuple[list[dict], int]:
        """查询采集数据"""
        db = self._get_db()
        conn = db.get_connection()
        
        conditions = ["1=1"]
        params = []
        
        if filters.get('merchantId'):
            conditions.append("merchant_id = ?")
            params.append(filters['merchantId'])
        
        if filters.get('id'):
            conditions.append("id = ?")
            params.append(filters['id'])

        if filters.get('ids'):
            id_list = filters['ids']
            if isinstance(id_list, list) and len(id_list) > 0:
                placeholders = ','.join(['?' for _ in id_list])
                conditions.append(f"id IN ({placeholders})")
                params.extend(id_list)

        where_clause = " AND ".join(conditions)
        offset = (page - 1) * page_size
        logger.info(f"数据查询条件: {where_clause} offset {offset} limit {page_size}")
        try:
            # 查询总数
            count_cursor = conn.execute(
                f"SELECT COUNT(*) as total FROM collected_data WHERE {where_clause}",
                tuple(params)
            )
            total = count_cursor.fetchone()['total']
            logger.info(f"Total collected_data: {total}")
            # 查询数据列表
            cursor = conn.execute(f"""
                SELECT * FROM collected_data 
                WHERE {where_clause}
                ORDER BY collected_at DESC
                LIMIT ? OFFSET ?
            """, (*params, page_size, offset))
            
            records = [DatabaseManager.dict_from_row(row) for row in cursor.fetchall()]
            
            # 反序列化 raw_data JSON
            for record in records:
                try:
                    record['raw_data'] = json.loads(record.get('raw_data', '{}'))
                except (json.JSONDecodeError, TypeError):
                    pass
            
            return records, total
        finally:
            conn.close()

    def delete(self, record_id: int) -> bool:
        """删除单条记录"""
        db = self._get_db()
        conn = db.get_connection()
        
        try:
            conn.execute("DELETE FROM collected_data WHERE id = ?", (record_id,))
            conn.commit()
            return True
        finally:
            conn.close()

    def export_to_excel(self, merchant_id: str = None, id: str = None, ids: list = None) -> str:
        """
        导出数据到 Excel 文件
        每行 = 一次任务执行结果
        列 = 固定信息列 + 字段映射后的动态字段列
        
        Args:
            merchant_id: 商家ID（可选）
            id: 单条记录ID（可选，兼容旧接口）
            ids: 批量导出的记录ID列表（可选）
        
        Returns:
            导出文件路径
        """
        import os
        import tempfile
        from datetime import datetime

        filters = {}
        if merchant_id:
            filters['merchantId'] = merchant_id
        if id:
            filters['id'] = id
        if ids:
            filters['ids'] = ids
        
        records, _ = self.list(filters, page=1, page_size=50000)
        
        if not records:
            raise Exception("没有可导出的数据")
        
        # 收集所有可能的列名（从所有记录的 raw_data 中取并集）
        # 过滤掉旧版废弃字段和元数据字段
        _EXCLUDE_KEYS = {'数据类型'}
        all_keys = set()
        rows_data = []
        for record in records:
            raw_data = record.get('raw_data', {})
            meta_info = {
                'id': record['id'],
                'task_id': record.get('task_id', ''),
                'merchant_name': record.get('merchant_name', ''),
                'collected_at': record['collected_at'],
            }

            if isinstance(raw_data, dict) and '_records' in raw_data and isinstance(raw_data['_records'], list):
                # 分页合并模式：每条提取记录展为独立一行
                for item in raw_data['_records']:
                    if not isinstance(item, dict):
                        continue
                    flat = {k: v for k, v in item.items() if not isinstance(v, (dict, list)) and k not in _EXCLUDE_KEYS}
                    all_keys.update(flat.keys())
                    rows_data.append({**meta_info, 'fields': flat})
            elif isinstance(raw_data, dict):
                # 单条记录模式：顶层字段作为列
                flat = {k: v for k, v in raw_data.items() if not k.startswith('_') and not isinstance(v, (dict, list)) and k not in _EXCLUDE_KEYS}
                all_keys.update(flat.keys())
                rows_data.append({**meta_info, 'fields': flat})
            else:
                all_keys.add('_raw')
                rows_data.append({**meta_info, 'fields': {'_raw': str(raw_data)}})
        
        # 固定列在前 + 动态字段列在后
        fixed_cols = ['序号', '商家名称', '采集时间']
        dynamic_cols = sorted(all_keys)
        headers = fixed_cols + dynamic_cols
        
        # 尝试使用 openpyxl 导出
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "采集数据"
            
            # 表头样式
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # 写表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # 填充数据行
            for row_idx, row_item in enumerate(rows_data, 2):
                fields = row_item['fields']
                
                # 固定列
                ws.cell(row=row_idx, column=1, value=row_idx - 1).border = border   # 序号
                ws.cell(row=row_idx, column=2, value=row_item['merchant_name']).border = border
                ts_val = datetime.fromtimestamp(row_item['collected_at']).strftime('%Y-%m-%d %H:%M:%S') if row_item['collected_at'] else ''
                ws.cell(row=row_idx, column=3, value=ts_val).border = border
                
                # 动态字段列（映射后的目标字段值）
                for col_idx, key in enumerate(dynamic_cols, len(fixed_cols) + 1):
                    val = fields.get(key, '')
                    if val is None:
                        val = ''
                    elif isinstance(val, (dict, list)):
                        val = json.dumps(val, ensure_ascii=False)
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # 自动调整列宽
            ws.column_dimensions['A'].width = 8       # 序号
            ws.column_dimensions['B'].width = 18      # 商家名称
            ws.column_dimensions['C'].width = 20      # 采集时间
            for col_idx, key in enumerate(dynamic_cols, 4):  # 从第4列开始（动态字段）
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                max_len = max(len(str(key)), *(len(str(r['fields'].get(key, ''))[:50]) for r in rows_data))
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)
            
            # 冻结首行
            ws.freeze_panes = 'A2'
            
            # 保存文件
            export_dir = os.path.join(os.getcwd(), 'exports')
            os.makedirs(export_dir, exist_ok=True)
            
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(export_dir, filename)
            
            wb.save(filepath)
            logger.info(f"Data exported to {filepath}: {len(rows_data)} rows, {len(headers)} cols ({fixed_cols} fixed + {dynamic_cols})")
            return filepath
            
        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV")
            return self._export_csv(rows_data, headers)
    
    def _export_csv(self, rows_data: list, headers: list) -> str:
        """CSV 格式导出（备用方案）"""
        import csv
        import os
        import json
        from datetime import datetime

        fixed_cols = ['序号', '商家名称', '采集时间']
        export_dir = os.path.join(os.getcwd(), 'exports')
        os.makedirs(export_dir, exist_ok=True)

        filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        filepath = os.path.join(export_dir, filename)

        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(headers)

            for row_idx, row_item in enumerate(rows_data, 1):
                fields = row_item['fields']
                ts_val = datetime.fromtimestamp(row_item['collected_at']).strftime('%Y-%m-%d %H:%M:%S') if row_item['collected_at'] else ''
                row = [row_idx, row_item['merchant_name'], ts_val]
                dynamic_headers = headers[len(fixed_cols):]
                for key in dynamic_headers:
                    val = fields.get(key, '')
                    if val is None:
                        val = ''
                    elif isinstance(val, (dict, list)):
                        val = json.dumps(val, ensure_ascii=False)
                    row.append(val)
                writer.writerow(row)

        return filepath
